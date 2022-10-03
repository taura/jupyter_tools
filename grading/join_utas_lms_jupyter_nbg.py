#!/usr/bin/python3

import argparse
import csv
import sys

import openpyxl

def load_worksheet(xlsx, sheet=0, header_row=None, start_row=None):
    wb = openpyxl.load_workbook(xlsx, data_only=True) # , read_only=True
    if isinstance(sheet, type(0)):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]
    rows = list(ws.rows)
    if header_row is None:
        header = None
    else:
        header = [cell.value for cell in rows[header_row]]
    if start_row is None:
        if header_row is None:
            start_row = 0
        else:
            start_row = header_row + 1
    return header, rows[start_row:]

def make_xls_rows(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    return ws.rows

def load_and_pivot_csv(a_csv, row_fields, col_fields, val_field):
    data = {}
    all_row_keys = set()
    all_col_keys = set()
    with open(a_csv) as csv_fp:
        fp = csv.DictReader(csv_fp)
        for row in fp:
            row_keys = tuple([row[field] for field in row_fields])
            col_keys = tuple([row[field] for field in col_fields])
            val = row[val_field]
            all_row_keys.add(row_keys)
            all_col_keys.add(col_keys)
            assert((row_keys + col_keys) not in data)
            data[row_keys + col_keys] = val
    rows = []
    all_row_keys = sorted(list(all_row_keys))
    all_col_keys = sorted(list(all_col_keys))
    for row_keys in all_row_keys:
        row = [row_keys[i] for i in range(len(row_fields))]
        for col_keys in all_col_keys:
            row.append(data.get(row_keys + col_keys, ""))
        rows.append(row)
    header = row_fields + ["-".join(col_keys) for col_keys in all_col_keys]
    return header, make_xls_rows(rows)

def merge_rows(rows0, make_key0, rows1, make_key1):
    rows0 = sorted(rows0, key=make_key0)
    rows1 = sorted(rows1, key=make_key1)
    n0 = len(rows0)
    n1 = len(rows1)
    ei = (None,) * len(rows0[0]) if n0 > 0 else ()
    ej = (None,) * len(rows1[0]) if n1 > 0 else ()
    merged_rows = []
    i = 0
    j = 0
    while i < n0 or j < n1:
        if j >= n1:
            # rows1 have ran out. fill rows1's columns with None
            ri = rows0[i]
            merged_rows.append(ri + ej)
            i += 1
        elif i >= n0:
            # rows0 have ran out. fill rows0's columns with None
            rj = rows1[j]
            merged_rows.append(ei + rj)
            j += 1
        else:
            ri = rows0[i]
            rj = rows1[j]
            ki = make_key0(ri)
            kj = make_key1(rj)
            if ki < kj:
                merged_rows.append(ri + ej)
                i += 1
            elif ki > kj:
                merged_rows.append(ei + rj)
                j += 1
            else:
                assert(ki == kj)
                merged_rows.append(ri + rj)
                i += 1
                j += 1
    return merged_rows

def cell_val(cell):
    if cell is None:
        return ""
    elif cell.value is None:
        return ""
    else:
        return cell.value

def make_joined_wb(utas_header, utas_rows, lms_header, lms_rows,
                   jupyter_header, jupyter_rows, nbg_header, nbg_rows):
    print("utas_header = {}".format(utas_header))
    utas_id_col = utas_header.index("学生証番号")
    def utas_key(row):
        return cell_val(row[utas_id_col])
    print("lms_header = {}".format(lms_header))
    lms_id_col = lms_header.index("学生証番号")
    def lms_key(row):
        return cell_val(row[lms_id_col]).replace("-", "")
    print("jupyter_header = {}".format(jupyter_header))
    # student id in juptyer.xlsx
    jupyter_id_col = jupyter_header.index("id")
    # symbolic user name in juptyer.xlsx (u22000, u22001, ...)
    jupyter_user_col = jupyter_header.index("user")
    # jupyter_user_col = jupyter_header.index("uid")
    def jupyter_id_key(row):
        return cell_val(row[jupyter_id_col]).replace("-", "")
    # join UTAL and LMS with UTAS 学生証番号 = LMS 学生証番号
    utas_lms_rows = merge_rows(utas_rows, utas_key, lms_rows, lms_key)
    utas_lms_n_cols = len(utas_lms_rows[0])
    # join UTAL+LMS and JUPYTER with UTAS/LMS 学生証番号 = JUPYTER id
    utas_lms_jupyter_rows = merge_rows(utas_lms_rows, utas_key,
                                       jupyter_rows, jupyter_id_key)
    def jupyter_user_key(row):
        return cell_val(row[utas_lms_n_cols + jupyter_user_col])
    # join UTAL+LMS+JUPYTER and NBG with JUPYTER UID = NBG STUDENT ID
    print("nbg_header = {}".format(nbg_header))
    nbg_student_id_col = nbg_header.index("student_id")
    def nbg_user_key(row):
        return cell_val(row[nbg_student_id_col])
    if 1:
        all_rows = merge_rows(utas_lms_jupyter_rows, jupyter_user_key,
                              nbg_rows, nbg_user_key)
        all_header = utas_header + lms_header + jupyter_header + nbg_header
    else:
        all_rows = utas_lms_jupyter_rows
        all_header = utas_header + lms_header + jupyter_header
    return all_header, all_rows

def save_xlsx(header, rows, a_xlsx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append([cell.value if cell is not None else None for cell in row])
    wb.save(a_xlsx)

def parse_args(argv):
    psr = argparse.ArgumentParser()
    psr.add_argument("--utas", metavar="UTAS_XLSX",
                     default="data/utas.xlsx",
                     help="UTAS Excel (UTAS -> 成績登録 -> select your lecture -> Excel出力 -> Excel (和文)); remove password by opening it with Windows Excel and File -> 情報 -> ブックの保護 -> パスワードによる暗号化")
    psr.add_argument("--lms", metavar="LMS_XLSX",
                     default="data/lms.xlsx",
                     help="LMS assignment Excel (ITC-LMS -> 課題 -> 全体提出状況確認 -> ダウンロード)")
    psr.add_argument("--jupyter", metavar="JUPYTER_XLSX",
                     default="data/jupyter.xlsx",
                     help="Jupyter Excel (Jupyter Googlesheet)")
    psr.add_argument("--nbg", metavar="NBG_XLSX",
                     default="data/nbg.xlsx",
                     help='nbgrader Excel (run work.py export to get grade.csv; save it as nbg.xlsx with Excel; open it with libreoffice and make a pivot table with row=student_id, column=assignment_name,notebook_name,prob_name, value=sum manual score; copy the contents to another sheet, add a row having a column name =concat(b2, "-", b3, "-", b4))')
    psr.add_argument("--grade", metavar="GRADE_CSV",
                     default="grade.csv",
                     help='grade.csv generated by work.py export')
    psr.add_argument("--out", metavar="UTAS_LMS_JUPYTER_NBGRADER_XLSX",
                     default="utas_lms_jupyter_nbgrader.xlsx",
                     help="Output Excel")
    args = psr.parse_args(argv[1:])
    return args

def main():
    args = parse_args(sys.argv)
    utas_header, utas_rows = load_worksheet(args.utas, header_row=3)
    lms_header, lms_rows = load_worksheet(args.lms, sheet='課題全体提出状況',
                                          header_row=1)
    jupyter_header, jupyter_rows = load_worksheet(args.jupyter, header_row=0)
    if 0:
        nbg_header, nbg_rows = load_worksheet(args.nbg, header_row=4)
    else:
        nbg_header, nbg_rows = load_and_pivot_csv(args.grade, ["student_id"], ["assignment_name", "notebook_name", "prob_name"], "manual_score")
    header, rows = make_joined_wb(utas_header, utas_rows, lms_header, lms_rows,
                                  jupyter_header, jupyter_rows, nbg_header, nbg_rows)
    save_xlsx(header, rows, args.out)
    return rows
    
main()


